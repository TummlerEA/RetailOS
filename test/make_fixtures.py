#!/usr/bin/env python3
"""
Rebuilds the .xlsx fixtures used by test_browser.js.

They are committed, so this only needs running when the fixtures should
change. It deliberately writes the awkward shapes a real target file
arrives in: months as formatted dates rather than text, conversion in
percentage points rather than fractions, the data on the second sheet, a
store that does not exist, and a cell someone typed a word into.

    pip install openpyxl && python3 test/make_fixtures.py
"""
import os
from datetime import date
from openpyxl import Workbook

HERE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
os.makedirs(HERE, exist_ok=True)


def stores():
    wb = Workbook()
    ws = wb.active
    ws.title = "Store list"
    ws.append(["Store No", "Store Name", "Manager", "Channel", "Country", "Status", "Opened", "Sqm"])
    rows = [
        ("S001", "Oxford Street",         "A. Kowalski", "Retail",     "United Kingdom", "Trading", date(2019, 3, 14), 420),
        ("S002", "Bicester Village",      "R. Mensah",   "Outlet",     "United Kingdom", "Trading", date(2021, 9, 2),  260),
        ("S003", "Selfridges Concession", "L. Ferrari",  "SIS",        "United Kingdom", "Trading", date(2023, 2, 20),  85),
        ("S004", "Trafford Centre",       "P. Nowak",    "Retail",     "United Kingdom", "Trading", date(2018, 6, 1),  310),
        ("S005", "retailos.example",      "—",           "Online",     "United Kingdom", "Trading", date(2016, 1, 1),  None),
    ]
    for row in rows:
        ws.append(row)
    for cell in ws["G"][1:]:
        cell.number_format = "dd/mm/yyyy"
    wb.save(os.path.join(HERE, "stores.xlsx"))


def targets_wide():
    """Months across the top, as real dates. Data on the second sheet."""
    wb = Workbook()
    cover = wb.active
    cover.title = "Read me"
    cover["A1"] = "FY26 operational targets — prepared by Planning"
    cover["A2"] = "Sales and traffic only. Sheet 'Targets' has the numbers."

    ws = wb.create_sheet("Targets")
    months = [date(2026, m, 1) for m in range(1, 7)]
    ws.append(["Store", "Store Name", "Metric"] + months)
    for cell in ws[1][3:]:
        cell.number_format = "mmm-yy"

    data = [
        ("S001", "Oxford Street",         "Sales",   [150000, 141000, 185000, 160000, 175000, 210000]),
        ("S001", "Oxford Street",         "Traffic", [23000, 22000, 26000, 24000, 25000, 29000]),
        ("S002", "Bicester Village",      "Sales",   [88000, 84000, 96000, 90000, 94000, 112000]),
        ("S002", "Bicester Village",      "Traffic", [17000, 16500, 19000, 18000, 18500, 21000]),
        ("S003", "Selfridges Concession", "Sales",   [40000, 38000, 44000, 41000, 43000, 51000]),
    ]
    for store, name, metric, values in data:
        ws.append([store, name, metric] + values)
    wb.save(os.path.join(HERE, "targets-wide.xlsx"))


def targets_long():
    """One row per store-month, conversion in percentage points, two bad rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Targets"
    ws.append(["Store ID", "Month", "Sales Target", "Traffic Target",
               "Conversion Target", "UPT Target", "ASP Target", "SOT Target"])
    rows = [
        ("S001", "Mar-26", 182000, 26000, 14.0, 1.85, 27.05, 7.00),
        ("S002", "Mar-26",  96000, 19000, 16.0, 2.10, 15.04, 5.05),
        ("S003", "Mar-26",  44000,  9000, 11.0, 1.60, 27.78, 4.89),
        ("S404", "Mar-26",  10000,  2000, 12.0, 1.50, 27.78, 5.00),   # store does not exist
        ("S004", "Mar-26", "tbc",  12000, 13.0, 1.70, 26.00, 4.00),   # a word where a number goes
    ]
    for row in rows:
        ws.append(row)
    wb.save(os.path.join(HERE, "targets-long.xlsx"))


stores()
targets_wide()
targets_long()
print("wrote fixtures to", HERE)
